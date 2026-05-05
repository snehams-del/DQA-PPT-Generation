from openpyxl import Workbook

from src.document_preprocessing.parser.xlsx_parser import _is_empty_row, parse


class TestIsEmptyRow:
    def test_all_none(self):
        assert _is_empty_row((None, None, None)) is True

    def test_all_empty_strings(self):
        assert _is_empty_row(("", "  ", "")) is True

    def test_mixed_none_and_empty(self):
        assert _is_empty_row((None, "", "  ")) is True

    def test_has_value(self):
        assert _is_empty_row((None, "data", None)) is False

    def test_has_numeric_value(self):
        assert _is_empty_row((None, 42, None)) is False

    def test_zero_is_not_empty(self):
        assert _is_empty_row((0,)) is False


class TestParse:
    def test_single_sheet(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        f = tmp_path / "test.xlsx"
        wb.save(str(f))
        wb.close()

        result = parse(str(f))

        assert "## Sheet1" in result
        assert "| Name | Age |" in result
        assert "| --- | --- |" in result
        assert "| Alice | 30 |" in result
        assert "| Bob | 25 |" in result

    def test_multiple_sheets(self, tmp_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.append(["X", "Y"])
        ws1.append([1, 2])

        ws2 = wb.create_sheet("Summary")
        ws2.append(["Total"])
        ws2.append([100])

        f = tmp_path / "multi.xlsx"
        wb.save(str(f))
        wb.close()

        result = parse(str(f))

        assert "## Data" in result
        assert "## Summary" in result
        assert "| Total |" in result

    def test_skips_empty_rows(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Header"])
        ws.append([None])
        ws.append(["Value"])
        f = tmp_path / "gaps.xlsx"
        wb.save(str(f))
        wb.close()

        result = parse(str(f))
        lines = [line for line in result.split("\n") if line.startswith("|")]
        data_lines = [line for line in lines if "---" not in line]
        assert len(data_lines) == 2

    def test_empty_sheet_skipped(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        f = tmp_path / "empty.xlsx"
        wb.save(str(f))
        wb.close()

        result = parse(str(f))
        assert result == ""

    def test_none_cells_rendered_as_empty(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", None, "C"])
        ws.append([1, 2, 3])
        f = tmp_path / "nulls.xlsx"
        wb.save(str(f))
        wb.close()

        result = parse(str(f))
        assert "| A |  | C |" in result
