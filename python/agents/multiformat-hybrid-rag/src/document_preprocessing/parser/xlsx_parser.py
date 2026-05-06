# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""XLSX → Markdown tables using openpyxl."""

from __future__ import annotations

import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _is_empty_row(row: tuple) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def parse(file_path: str, genai_client=None, max_workers: int = 8) -> str:
    """Extract markdown tables from a .xlsx file."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [row for row in ws.iter_rows(values_only=True) if not _is_empty_row(row)]

        if not rows:
            continue

        parts.append(f"## {sheet_name}")

        table_rows = []
        for row in rows:
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            table_rows.append("| " + " | ".join(cells) + " |")

        if table_rows:
            num_cols = len(rows[0])
            header_sep = "| " + " | ".join(["---"] * num_cols) + " |"
            table_rows.insert(1, header_sep)
            parts.append("\n".join(table_rows))

    wb.close()
    return "\n\n".join(parts)
